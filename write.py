import numpy as np
import openpyxl as xl
from openpyxl.utils.cell import column_index_from_string
from datetime import datetime

from util import addYears, decaler


def use_scenario(sheet_pft, scenario, dico_nd):
    first_line = 13
    last_line = sheet_pft.max_row
    # today = datetime.now()
    # Date des données, je ne décale que les jalons présents après cette date
    today = datetime.strptime("01/01/23 10:00:00", "%d/%m/%y %H:%M:%S")
    min_do = datetime.strptime("01/01/23 10:00:00", "%d/%m/%y %H:%M:%S")
    max_do = datetime.strptime("31/12/32 10:00:00", "%d/%m/%y %H:%M:%S")

    col_typo = "C"
    pos_typo = column_index_from_string(col_typo)
    col_statut = "F"
    pos_statut = column_index_from_string(col_statut)
    col_moa = "DM"
    pos_moa = column_index_from_string(col_moa)
    col_sp = "DR"
    pos_sp = column_index_from_string(col_sp)

    # cotation retenue
    col_prio = "IK"
    pos_prio = column_index_from_string(col_prio)

    # cotation forcée
    col_prio_inter = "IL"
    pos_prio_inter = int(column_index_from_string(col_prio_inter))

    # anticipation ou retard forcé
    col_decalage_inter = "IM"
    pos_decalage_inter = int(column_index_from_string(col_decalage_inter))

    # colonne de l'année 2023
    col_consistances = "DU"
    pos_consistances = int(column_index_from_string(col_consistances))

    col_ressources_bleues = "JB"
    pos_ressources_bleues = int(column_index_from_string(col_ressources_bleues))

    col_probable = "MA"
    pos_probable = int(column_index_from_string(col_probable))

    col_jalons = "IU"
    pos_jalons = int(column_index_from_string(col_jalons))

    for line in range(first_line, last_line + 1):
        moa = str(sheet_pft.cell(line, pos_moa).value)
        sp = str(sheet_pft.cell(line, pos_sp).value)
        statut = str(sheet_pft.cell(line, pos_statut).value)
        typo = str(sheet_pft.cell(line, pos_typo).value)

        couple = f"{moa}-{sp}"
        key = None
        if moa in scenario:
            key = moa
        elif couple in scenario:
            key = couple

        # même si ma clé est vide, je force le décalage si un décalage inter est renseigné
        if sheet_pft.cell(line, pos_decalage_inter).value is None:
            decalage_inter_is_filled = False
        else:
            decalage_inter_is_filled = True

        if key is not None or decalage_inter_is_filled:

            # si la ligne du scénario n'est pas ND ou si la colonne B du portefeuille est Nd, je prends
            if key not in dico_nd or typo == "ND" or decalage_inter_is_filled:
                # si la pos_prio_inter est vide, je prends la prio classique
                if sheet_pft.cell(line, pos_prio_inter).value is None:
                    prio = str(sheet_pft.cell(line, pos_prio).value)
                else:
                    prio = str(sheet_pft.cell(line, pos_prio_inter).value)

                if not prio.isdigit() or prio == "0":
                    prio = "9"

                # j'applique le scénario à cette ligne
                # si la colonne pos_decalage_inter est vide, je prends le décalage classique
                if sheet_pft.cell(line, pos_decalage_inter).value is None:
                    decalage = int(scenario[key][prio])
                else:
                    decalage = int(sheet_pft.cell(line, pos_decalage_inter).value)

                valeur_do = sheet_pft.cell(line, pos_jalons).value
                if decalage > 0 and addYears(valeur_do, decalage) > max_do:
                    sheet_pft.cell(line, 1).value = "retard trop important"
                    continue

                elif decalage < 0 and addYears(valeur_do, decalage) < min_do:
                    sheet_pft.cell(line, 1).value = "anticipation trop importante"
                    continue

                else:
                    sheet_pft.cell(line, 1).value = decalage

                # consistances
                for num_consistance in range(10):
                    nom_consistance = sheet_pft.cell(line, pos_consistances + 12 * num_consistance - 2).value
                    if nom_consistance == "" or nom_consistance is None:
                        continue
                    else:
                        list_valeurs = []
                        for annee in range(10):
                            # stockage des valeurs présentes
                            list_valeurs.append(sheet_pft.cell(line, pos_consistances + 12 * num_consistance + annee).value or 0)
                            # suppression des valeurs
                            sheet_pft.cell(line, pos_consistances + 12 * num_consistance + annee).value = None

                        list_valeurs_decalees = decaler(list_valeurs, decalage)
                        # recopie de la consistance avec le décalage
                        for annee in range(10):
                            sheet_pft.cell(line, pos_consistances + 12 * num_consistance + annee).value = list_valeurs_decalees[annee].replace(",", ".")

                # Ressources
                for num_ressources in range(7):
                    list_valeurs = []
                    for annee in range(10):
                        # stockage des valeurs présentes
                        list_valeurs.append(sheet_pft.cell(line, pos_ressources_bleues + 12 * num_ressources + annee).value or 0)
                        # suppression des valeurs
                        sheet_pft.cell(line, pos_ressources_bleues + 12 * num_ressources + annee).value = None

                    list_valeurs_decalees = decaler(list_valeurs, decalage)
                    # recopie des ressources avec le décalage
                    for annee in range(10):
                        sheet_pft.cell(line, pos_ressources_bleues + 12 * num_ressources + annee).value = list_valeurs_decalees[annee].replace(",", ".")

                # Probable
                list_valeurs = []
                for annee in range(10):
                    # stockage des valeurs présentes
                    list_valeurs.append(sheet_pft.cell(line, pos_probable + annee).value or 0)
                    # suppression des valeurs
                    sheet_pft.cell(line, pos_probable + annee).value = None

                list_valeurs_decalees = decaler(list_valeurs, decalage)
                # recopie du probable avec le décalage
                for annee in range(0, 10):
                    sheet_pft.cell(line, pos_probable + annee).value = list_valeurs_decalees[annee].replace(",", ".")

                # Jalons
                if not (statut == "Jalons manquants" or statut == "---"):
                    for jalon in range(6):
                        valeur_pft = sheet_pft.cell(line, pos_jalons + jalon).value

                        if str(valeur_pft) == "0" or str(valeur_pft) == "":
                            sheet_pft.cell(line, pos_probable + jalon + decalage).value = None
                        elif isinstance(valeur_pft, str):
                            try:
                                valeur_pft = datetime.strptime(valeur_pft, '%d/%m/%Y')
                            except:
                                break
                            if valeur_pft > today:
                                # je décale seulement les jalons dans le futur
                                new_date = addYears(valeur_pft, decalage)
                                sheet_pft.cell(line, pos_jalons + jalon).value = new_date
                        elif valeur_pft > today:
                            # je décale seulement les jalons dans le futur
                            new_date = addYears(valeur_pft, decalage)
                            sheet_pft.cell(line, pos_jalons + jalon).value = new_date
                        else:
                            pass
        else:
            # pas de décalage, je ne fais rien
            pass
